"""
Excel file creation service.
Uses openpyxl to create .xlsx files with headers, data rows, and formatting.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from shared.logger import get_logger

log = get_logger("excel_service")

_DEFAULT_DIR = Path.home() / "Desktop"


class ExcelService:

    async def create_excel(
        self,
        filename: str,
        sheets: List[Dict[str, Any]],
        save_dir: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Create an Excel workbook with one or more sheets.

        Args:
            filename: Output filename (must end in .xlsx)
            sheets: List of sheet dicts, each with:
                - name: str (sheet tab name)
                - headers: list[str] (column headers)
                - rows: list[list] (data rows)
                - column_widths: optional dict of col_letter -> width
            save_dir: Directory to save the file (default: Desktop)

        Returns:
            {success, filepath, error}
        """
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"

        save_path = Path(save_dir or _DEFAULT_DIR) / filename
        save_path.parent.mkdir(parents=True, exist_ok=True)

        try:
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)

            for sheet_data in sheets:
                name = sheet_data.get("name", "Sheet1")[:31]  # Excel limit
                headers = sheet_data.get("headers", [])
                rows = sheet_data.get("rows", [])
                column_widths = sheet_data.get("column_widths", {})

                ws = wb.create_sheet(title=name)

                # Write headers
                if headers:
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_idx, value=header)
                        cell.font = Font(bold=True, color="FFFFFF", size=11)
                        cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = Border(
                            left=Side(style="thin"),
                            right=Side(style="thin"),
                            top=Side(style="thin"),
                            bottom=Side(style="thin"),
                        )

                # Write data rows
                for row_idx, row_data in enumerate(rows, start=2):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.alignment = Alignment(vertical="center")
                        cell.border = Border(
                            left=Side(style="thin"),
                            right=Side(style="thin"),
                            top=Side(style="thin"),
                            bottom=Side(style="thin"),
                        )

                # Auto-adjust column widths if not specified
                if headers:
                    for col_idx in range(1, len(headers) + 1):
                        col_letter = get_column_letter(col_idx)
                        if col_letter not in column_widths:
                            max_len = len(str(headers[col_idx - 1]))
                            for row in rows:
                                if col_idx <= len(row):
                                    val_len = len(str(row[col_idx - 1]))
                                    max_len = max(max_len, val_len)
                            ws.column_dimensions[col_letter].width = min(max_len + 3, 60)
                        else:
                            ws.column_dimensions[col_letter].width = column_widths[col_letter]

                # Freeze top row
                ws.freeze_panes = "A2"

                # Auto-filter
                if headers:
                    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

            wb.save(str(save_path))
            log.info("excel_created", path=str(save_path), sheets=len(sheets))

            return {
                "success": True,
                "filepath": str(save_path),
                "filename": filename,
                "sheets": len(sheets),
                "rows": sum(len(s.get("rows", [])) for s in sheets),
            }

        except Exception as e:
            log.error("excel_creation_failed", error=str(e))
            return {"success": False, "filepath": "", "error": str(e)}

    async def create_from_csv_data(
        self,
        filename: str,
        sheet_name: str,
        headers: List[str],
        rows: List[List[Any]],
        save_dir: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Quick helper to create a single-sheet Excel from headers+rows."""
        return await self.create_excel(
            filename=filename,
            sheets=[{
                "name": sheet_name,
                "headers": headers,
                "rows": rows,
            }],
            save_dir=save_dir,
        )


excel_service = ExcelService()
